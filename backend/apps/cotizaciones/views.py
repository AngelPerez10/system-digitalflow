"""ViewSets for cotizaciones app."""
import io
import logging
from types import SimpleNamespace

from django.db import models as django_models
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image as PILImage
from rest_framework import filters, viewsets
from rest_framework.decorators import action
from rest_framework.response import Response

from apps.common.pdf_html import subtotal_iva_display_split as _subtotal_iva_display_split
from apps.common.pdf_images import safe_http_image_bytes as _safe_http_image_bytes
from apps.users.permissions import ModulePermission, user_module_own_only

from .categorias_productos import categorias_nombres_por_id, normalize_categorias_productos
from .models import Cotizacion
from .pdf_render import PdfRenderError, any_provider_configured, render_html_to_pdf
from .serializers import CotizacionSerializer

logger = logging.getLogger(__name__)

IVA_MX_DISPLAY = 1.16
ANTICIPO_PCT = 60


def _cotizacion_item_line_totals(
    cantidad: float,
    precio_lista: float,
    descuento_pct: float,
    producto_externo_id: str,
) -> tuple[float, float, float]:
    """
    Importes por línea (misma regla que CotizacionSerializer._calculate_totals).
    Retorna (precio_unitario_con_iva, importe_con_iva, importe_sin_iva).
    """
    producto_externo_id = str(producto_externo_id or "").strip()
    es_manual = producto_externo_id == ""
    descuento_pct = float(descuento_pct or 0)
    if descuento_pct < 0:
        descuento_pct = 0.0
    if descuento_pct > 100:
        descuento_pct = 100.0

    factor = 1.0 - (descuento_pct / 100.0)
    pu_lista_desc = float(precio_lista or 0) * factor
    if es_manual:
        pu_sin_iva = pu_lista_desc
        pu_con_iva = pu_lista_desc * IVA_MX_DISPLAY
    else:
        pu_sin_iva = (float(precio_lista or 0) / IVA_MX_DISPLAY) * factor
        pu_con_iva = pu_lista_desc

    qty = float(cantidad or 0)
    importe_con_iva = qty * pu_con_iva
    importe_sin_iva = qty * pu_sin_iva
    return pu_con_iva, importe_con_iva, importe_sin_iva


def _user_display(user) -> str:
    if not user:
        return "—"
    try:
        first = getattr(user, "first_name", "") or ""
        last = getattr(user, "last_name", "") or ""
        full = f"{first} {last}".strip()
        if full:
            return full
        return getattr(user, "username", None) or getattr(user, "email", None) or str(getattr(user, "pk", ""))
    except Exception:
        logger.exception("Failed to resolve user display name")
        return "—"


def _tipo_trabajo_labels(cotizacion: Cotizacion) -> str:
    try:
        servicios = cotizacion.tipo_trabajo.all()
    except Exception:
        logger.exception("Failed to load tipo_trabajo for export")
        return "—"
    nombres = [str(getattr(s, "nombre", "") or "").strip() for s in servicios]
    nombres = [n for n in nombres if n]
    return ", ".join(nombres) if nombres else "—"


def _medio_label(raw: str) -> str:
    s = str(raw or "").strip().upper().replace(" ", "_")
    m = {
        "BNI": "BNI",
        "REFERIDO": "Referido",
        "WEB": "Web",
        "TIENDA_ONLINE": "Tienda Online",
        "FACEBOOK": "Facebook",
        "INSTAGRAM": "Instagram",
        "TIKTOK": "Tiktok",
        "GOOGLE_MAPS": "Google Maps",
        "YOUTUBE": "Youtube",
        "TIENDA_FISICA": "Tienda Fisica",
        "OTRO": "Otro",
    }
    return m.get(s, str(raw or "").strip() or "—")


def _status_label(raw: str) -> str:
    s = str(raw or "").strip().upper()
    if s == "AUTORIZADA":
        return "Autorizada"
    if s == "CANCELADA":
        return "Cancelada"
    if s == "PENDIENTE" or not s:
        return "Pendiente"
    return str(raw or "").strip() or "—"


COTIZACION_EXPORT_BRAND_HEX = "3160E3"


def _build_cotizacion_excel_bytes(cotizacion: Cotizacion) -> bytes:
    """Genera un .xlsx con encabezado, líneas (con miniatura) y totales."""
    from .pdf_opciones import parse_pdf_opciones_from_cotizacion

    pdf_opciones = parse_pdf_opciones_from_cotizacion(cotizacion)
    show_detalle = not pdf_opciones.ocultar_detalle
    wb = Workbook()
    ws = wb.active
    ws.title = "Cotización"

    title_font = Font(bold=True, color="FFFFFF", size=13)
    title_fill = PatternFill("solid", fgColor=COTIZACION_EXPORT_BRAND_HEX)
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill("solid", fgColor=COTIZACION_EXPORT_BRAND_HEX)
    label_font = Font(bold=True, color="111827", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    center_wrap = Alignment(wrap_text=True, horizontal="center", vertical="center")
    thin = Side(style="thin", color="D1D5DB")
    table_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    zebra_fill = PatternFill("solid", fgColor="F9FAFB")
    label_fill = PatternFill("solid", fgColor="F3F4F6")
    value_fill = PatternFill("solid", fgColor="FFFFFF")

    cliente_obj = getattr(cotizacion, "cliente_id", None)
    if cliente_obj is not None and not isinstance(cliente_obj, django_models.Model):
        try:
            from apps.clientes.models import Cliente

            cliente_obj = Cliente.objects.filter(id=cliente_obj).first() or None
        except Exception:
            logger.exception("Failed to load Cliente for Excel export")
            cliente_obj = None

    cliente_nombre = (getattr(cliente_obj, "nombre", None) or cotizacion.cliente or "").strip() or "—"
    folio = getattr(cotizacion, "idx", None) or cotizacion.id
    fecha = cotizacion.fecha.strftime("%d/%m/%Y") if cotizacion.fecha else "—"

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    title_cell = ws.cell(row=r, column=1, value="COTIZACIÓN")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = table_border
    ws.row_dimensions[r].height = 28
    r += 2

    meta_pairs = [
        ("Folio", folio),
        ("Fecha", fecha),
        ("Medio de contacto", _medio_label(getattr(cotizacion, "medio_contacto", "") or "")),
        ("Status", _status_label(getattr(cotizacion, "status", "") or "")),
        ("Creada por", _user_display(getattr(cotizacion, "creado_por", None))),
        ("Editada por", _user_display(getattr(cotizacion, "actualizado_por", None))),
        ("Cliente", cliente_nombre),
        ("Contacto", str(cotizacion.contacto or "").strip() or "—"),
        ("Teléfono de contacto", str(getattr(cotizacion, "contacto_telefono", "") or "").strip() or "—"),
        ("Tipo de trabajo", _tipo_trabajo_labels(cotizacion)),
    ]
    try:
        dcp = float(cotizacion.descuento_cliente_pct or 0)
    except Exception:
        logger.exception("Failed to parse descuento_cliente_pct for Excel")
        dcp = 0.0
    meta_pairs.extend(
        [
            ("Descuento de Cliente (%)", f"{dcp:.2f}"),
            ("Moneda", "MXN"),
        ]
    )
    try:
        fc = cotizacion.fecha_creacion.strftime("%d/%m/%Y %H:%M") if getattr(cotizacion, "fecha_creacion", None) else "—"
    except Exception:
        logger.exception("Failed to format fecha_creacion")
        fc = "—"
    try:
        fa = cotizacion.fecha_actualizacion.strftime("%d/%m/%Y %H:%M") if getattr(cotizacion, "fecha_actualizacion", None) else "—"
    except Exception:
        logger.exception("Failed to format fecha_actualizacion")
        fa = "—"
    meta_pairs.extend(
        [
            ("Fecha creación", fc),
            ("Fecha última actualización", fa),
        ]
    )

    # Bloque horizontal: etiqueta/valor de izquierda a derecha (4 pares por fila)
    pair_cols = [(1, 2), (3, 4), (5, 6), (7, 8)]
    block_row = r
    for i, (label, value) in enumerate(meta_pairs):
        pair_idx = i % len(pair_cols)
        if pair_idx == 0 and i > 0:
            block_row += 1
        c_label, c_value = pair_cols[pair_idx]

        cl = ws.cell(row=block_row, column=c_label, value=label)
        cv = ws.cell(row=block_row, column=c_value, value=value)
        cl.font = label_font
        cl.fill = label_fill
        cl.alignment = Alignment(vertical="center")
        cl.border = table_border
        cv.fill = value_fill
        cv.alignment = wrap
        cv.border = table_border
    # Bordes suaves en columnas no usadas (9-10)
    for rr in range(r, block_row + 1):
        for cc in (9, 10):
            ws.cell(row=rr, column=cc).border = table_border
            ws.cell(row=rr, column=cc).fill = value_fill

    r = block_row + 2

    # Tabla de conceptos
    headers = [
        "Imagen",
        "Cantidad",
        "Unidad",
        "Descripción",
        "Detalle",
        "Precio unitario",
        "Descuento (%)",
        "Importe total",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_wrap
        cell.border = table_border
    ws.row_dimensions[r].height = 24
    r += 1
    first_data_row = r

    items_rel = getattr(cotizacion, "items", None)
    if items_rel is None:
        items = []
    elif hasattr(items_rel, "all"):
        try:
            items = list(items_rel.all())
        except Exception:
            logger.exception("Failed to load cotizacion items for Excel export")
            items = []
    elif isinstance(items_rel, (list, tuple)):
        items = list(items_rel)
    else:
        items = []

    cat_names = categorias_nombres_por_id(
        normalize_categorias_productos(getattr(cotizacion, "categorias_productos", None))
    )
    last_cat_id = None
    net_subtotal_con_iva = 0.0
    for it in items:
        cat_id = str(getattr(it, "categoria_id", "") or "").strip()
        if cat_id and cat_id in cat_names and cat_id != last_cat_id:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
            cat_cell = ws.cell(row=r, column=1, value=str(cat_names[cat_id]))
            cat_cell.font = Font(bold=True, color="3160E3", size=11)
            cat_cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[r].height = 22
            r += 1
            last_cat_id = cat_id

        try:
            cantidad = float(it.cantidad or 0)
            precio_lista = float(it.precio_lista or 0)
            descuento = float(it.descuento_pct or 0)
            producto_externo_id = str(getattr(it, "producto_externo_id", "") or "").strip()
            pu, importe, _importe_sin_iva = _cotizacion_item_line_totals(
                cantidad, precio_lista, descuento, producto_externo_id
            )
            net_subtotal_con_iva += importe
        except Exception:
            logger.exception("Failed to parse cotizacion item values for Excel")
            cantidad = 0.0
            precio_lista = 0.0
            descuento = 0.0
            pu = 0.0
            importe = 0.0

        ws.row_dimensions[r].height = 64
        ws.cell(row=r, column=2, value=cantidad)
        ws.cell(row=r, column=3, value=str(getattr(it, "unidad", "") or ""))
        nombre_base = str(getattr(it, "producto_nombre", "") or "")
        descripcion_larga = str(getattr(it, "producto_descripcion", "") or "")
        if pdf_opciones.simplificar_descripcion:
            corta = str(getattr(it, "pdf_descripcion_corta", "") or "").strip()
            if not corta:
                larga = descripcion_larga.strip()
                corta = (
                    larga[:120].rstrip() + ("…" if len(larga) > 120 else "")
                    if larga
                    else ""
                )
            producto_val = corta or nombre_base
            detalle_val = descripcion_larga if show_detalle else ""
        else:
            producto_val = nombre_base
            detalle_val = descripcion_larga if show_detalle else ""
        ws.cell(row=r, column=4, value=producto_val).alignment = wrap
        ws.cell(row=r, column=5, value=detalle_val).alignment = wrap
        ws.cell(row=r, column=6, value=round(pu, 2))
        ws.cell(row=r, column=7, value=round(descuento, 2))
        ws.cell(row=r, column=8, value=round(importe, 2))
        if (r - first_data_row) % 2 == 1:
            for cidx in range(1, 9):
                ws.cell(row=r, column=cidx).fill = zebra_fill
        for cidx in range(1, 9):
            ws.cell(row=r, column=cidx).border = table_border

        thumb = str(getattr(it, "thumbnail_url", "") or "").strip()
        if thumb:
            raw = _safe_http_image_bytes(thumb)
            if raw:
                try:
                    with PILImage.open(io.BytesIO(raw)) as pil_im:
                        out = io.BytesIO()
                        pil_im.save(out, format="PNG")
                        out.seek(0)
                    img = XLImage(out)
                    img.width = 54
                    img.height = 54
                    ws.add_image(img, f"A{r}")
                except Exception:
                    logger.exception("Failed to embed thumbnail image in Excel")

        r += 1

    last_data_row = max(first_data_row, r - 1)

    # Totales
    subtotal_lineas = net_subtotal_con_iva if net_subtotal_con_iva else float(cotizacion.subtotal or 0)
    if subtotal_lineas < 0:
        subtotal_lineas = 0.0

    descuento_cliente_pct = dcp
    if descuento_cliente_pct < 0:
        descuento_cliente_pct = 0.0
    if descuento_cliente_pct > 100:
        descuento_cliente_pct = 100.0

    descuento_cliente_monto = subtotal_lineas * (descuento_cliente_pct / 100.0)
    total_con_iva = max(0.0, subtotal_lineas - descuento_cliente_monto)
    total_guardado = float(cotizacion.total or 0)
    if total_guardado > 0:
        total_con_iva = total_guardado
    base_sin_iva, iva_display = _subtotal_iva_display_split(total_con_iva)

    money_fmt = '$#,##0.00'
    totals_label_font = Font(bold=True, color="111827")
    totals_rows_start = r + 1

    r = totals_rows_start
    ws.cell(row=r, column=7, value="Subtotal").font = totals_label_font
    ws.cell(row=r, column=8, value=round(base_sin_iva, 2))
    ws.cell(row=r, column=8).number_format = money_fmt

    r += 1
    ws.cell(row=r, column=7, value="IVA (16%)").font = totals_label_font
    ws.cell(row=r, column=8, value=round(iva_display, 2))
    ws.cell(row=r, column=8).number_format = money_fmt

    r += 1
    ws.cell(row=r, column=7, value="Total").font = Font(bold=True, color="FFFFFF")
    ws.cell(row=r, column=7).fill = PatternFill("solid", fgColor=COTIZACION_EXPORT_BRAND_HEX)
    ws.cell(row=r, column=8, value=round(total_con_iva, 2))
    ws.cell(row=r, column=8).number_format = money_fmt
    ws.cell(row=r, column=8).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=r, column=8).fill = PatternFill("solid", fgColor=COTIZACION_EXPORT_BRAND_HEX)

    for rr in range(totals_rows_start, r + 1):
        ws.cell(row=rr, column=7).border = table_border
        ws.cell(row=rr, column=8).border = table_border

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 34
    if not show_detalle:
        ws.column_dimensions["E"].hidden = True
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["J"].width = 12

    for rr in range(first_data_row, last_data_row + 1):
        for cc in range(2, 9):
            ws.cell(row=rr, column=cc).alignment = wrap
        ws.cell(row=rr, column=2).alignment = center_wrap
        ws.cell(row=rr, column=3).alignment = center_wrap
        ws.cell(row=rr, column=6).number_format = '$#,##0.00'
        ws.cell(row=rr, column=7).number_format = '0.00'
        ws.cell(row=rr, column=8).number_format = '$#,##0.00'

    ws.freeze_panes = f"A{first_data_row}"

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class CotizacionesPermission(ModulePermission):
    """Permission class for cotizaciones module."""
    module_key = 'cotizaciones'


class CotizacionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing Cotizacion instances.

    Provides CRUD operations for quotations with permission-based access control.
    Supports both client-based and prospect-based quotations.
    """
    queryset = Cotizacion.objects.all()
    serializer_class = CotizacionSerializer
    permission_classes = [CotizacionesPermission]
    pagination_class = None

    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = [
        'idx',
        'cliente',
        'contacto',
        'creado_por__username',
        'creado_por__first_name',
        'creado_por__last_name',
    ]
    ordering_fields = ['idx', 'fecha', 'medio_contacto', 'status', 'fecha_creacion', 'total']
    ordering = ['-idx']

    def get_queryset(self):
        """
        Get optimized queryset with prefetched items and related objects.

        Returns:
            QuerySet: Cotizacion queryset ordered by idx descending
        """
        queryset = super().get_queryset()
        queryset = queryset.prefetch_related(
            Prefetch(
                'items',
                queryset=Cotizacion.items.rel.related_model.objects.order_by('orden')
            ),
            'tipo_trabajo',
        )
        queryset = queryset.select_related('cliente_id', 'creado_por', 'actualizado_por')
        user = getattr(self.request, 'user', None)
        if user and getattr(user, 'is_authenticated', False):
            own_only = user_module_own_only(user, 'cotizaciones')
            if own_only:
                queryset = queryset.filter(Q(creado_por=user) | Q(actualizado_por=user))
        return queryset.order_by('-idx')

    def _generate_pdf_html(self, cotizacion: Cotizacion, pdf_opciones=None) -> str:
        from .pdf_templates import generate_cotizacion_pdf_html

        return generate_cotizacion_pdf_html(cotizacion, pdf_opciones=pdf_opciones)

    @action(
        detail=True,
        methods=['get'],
        url_path='pdf',
    )
    def pdf(self, request, pk=None):
        from .pdf_opciones import parse_pdf_opciones_from_cotizacion

        cotizacion = self.get_object()
        pdf_opciones = parse_pdf_opciones_from_cotizacion(cotizacion)

        html = self._generate_pdf_html(cotizacion, pdf_opciones=pdf_opciones)
        if not html:
            return Response({"detail": "No se pudo generar el HTML del PDF."}, status=500)

        idx = getattr(cotizacion, "idx", None) or cotizacion.id
        filename = f"Cotizacion_{idx}.pdf"

        # Allow clients to explicitly ask for HTML as a printable fallback
        # when the PDF engines are unavailable (e.g. ?format=html).
        wants_html = (request.query_params.get("format") or "").lower() == "html"
        if wants_html or not any_provider_configured():
            response = HttpResponse(html, content_type="text/html; charset=utf-8")
            if wants_html:
                response["Content-Disposition"] = f'inline; filename="Cotizacion_{idx}.html"'
            return response

        try:
            pdf_bytes = render_html_to_pdf(html, size="A4", landscape=False, timeout=90)
        except PdfRenderError as e:
            logger.exception("PDF render failed for cotizacion %s: %s", pk, e.detail)
            return Response({"detail": "No se pudo generar el PDF."}, status=502)

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = f'inline; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'], url_path='excel')
    def excel(self, request, pk=None):
        cotizacion = self.get_object()
        try:
            xlsx_bytes = _build_cotizacion_excel_bytes(cotizacion)
        except Exception:
            logger.exception("Excel generation failed for cotizacion %s", getattr(cotizacion, 'pk', '?'))
            return Response({"detail": "No se pudo generar el Excel."}, status=500)

        idx = getattr(cotizacion, "idx", None) or cotizacion.id
        filename = f"Cotizacion_{idx}.xlsx"
        response = HttpResponse(
            xlsx_bytes,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['post'], url_path='pdf-preview')
    def pdf_preview(self, request):
        from .pdf_opciones import parse_pdf_opciones_from_request_data

        pdf_opciones = parse_pdf_opciones_from_request_data(request.data)

        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        data = serializer.validated_data
        items_data = data.pop('items', [])
        raw_items = request.data.get('items') if isinstance(request.data, dict) else []
        if not isinstance(raw_items, list):
            raw_items = []

        items = []
        for idx, it in enumerate(items_data or []):
            raw_it = raw_items[idx] if idx < len(raw_items) and isinstance(raw_items[idx], dict) else {}
            item = SimpleNamespace(
                producto_externo_id=it.get('producto_externo_id'),
                cantidad=it.get('cantidad'),
                precio_lista=it.get('precio_lista'),
                descuento_pct=it.get('descuento_pct'),
                unidad=it.get('unidad'),
                producto_nombre=it.get('producto_nombre'),
                producto_descripcion=it.get('producto_descripcion'),
                thumbnail_url=it.get('thumbnail_url'),
                pdf_descripcion_corta=raw_it.get('pdf_descripcion_corta', ''),
                categoria_id=it.get('categoria_id', ''),
            )
            items.append(item)

        preview = SimpleNamespace(
            id=0,
            idx='PREVIEW',
            cliente_id=data.get('cliente_id'),
            cliente=data.get('cliente', ''),
            contacto=data.get('contacto', ''),
            contacto_telefono=data.get('contacto_telefono', ''),
            fecha=data.get('fecha'),
            vencimiento=data.get('vencimiento'),
            subtotal=data.get('subtotal', 0),
            descuento_cliente_pct=data.get('descuento_cliente_pct', 0),
            iva_pct=data.get('iva_pct', 0),
            iva=data.get('iva', 0),
            total=data.get('total', 0),
            texto_arriba_precios=data.get('texto_arriba_precios', ''),
            terminos=data.get('terminos', ''),
            categorias_productos=data.get('categorias_productos', []),
            items=items,
        )

        html = self._generate_pdf_html(preview, pdf_opciones=pdf_opciones)
        if not html:
            return Response({"detail": "No se pudo generar el HTML del PDF."}, status=500)

        if not any_provider_configured():
            return HttpResponse(html, content_type="text/html; charset=utf-8")

        try:
            pdf_bytes = render_html_to_pdf(html, size="A4", landscape=False, timeout=90)
        except PdfRenderError as e:
            logger.exception("PDF preview render failed: %s", e.detail)
            return Response({"detail": "No se pudo generar el PDF."}, status=502)

        response = HttpResponse(pdf_bytes, content_type="application/pdf")
        response["Content-Disposition"] = 'inline; filename="Cotizacion_Preview.pdf"'
        return response
